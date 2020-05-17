import requests
import csv
import glob
import os
import yaml
from openpyxl import load_workbook
from datetime import datetime
from gmailer import emailer

def load_yaml(file_path):
    with open(file_path) as file:
        file_object = yaml.safe_load(file)
        
    return file_object
    
def get_json(url):
    resp = requests.get(url)
    prediction = resp.json()
    return prediction
    
def fetch_json_data(json_body):
    json_data = []
    for entry in json_body:
        row = [entry]
        for k, v in json_body[entry].items():
            row.append(v)
            
        json_data.append(row)
    
    json_data.reverse()
    
    return json_data
    
def write_data(filepath, data, headers = []):
    with open(filepath, 'w', newline='') as file:
        writer = csv.writer(file)
        if len(headers) > 0:
            writer.writerow(headers)
        writer.writerows(data)
        
def clear_folder(folder_path):
    files = glob.glob(folder_path + '*')
    for f in files:
        os.remove(f)
        
def gather_api_stock_data(ticker, key):
    print("Gathering data for " + ticker)
    url = 'https://www.alphavantage.co/query?function=TIME_SERIES_DAILY_ADJUSTED&symbol=' + ticker + \
          '&outputsize=full&apikey=' + key
    json_body = get_json(url)['Time Series (Daily)']
    json_data = fetch_json_data(json_body)
    
    return json_data
    
def write_stock_data_to_workbook(data_path, workbook, ticker):
    print("Writing " + ticker + " data to workbook")
    workbook.create_sheet(ticker)
    sheet = workbook[ticker] 
    with open(data_path) as f:
        reader = csv.reader(f)
        sheet.append(next(reader))
        for row in reader:
            dtype_row = [datetime.strptime(row[0], '%Y-%m-%d')]
            for val in row[1:]:
                dtype_row.append(float(val))
            sheet.append(dtype_row)
            
def write_stock_names_to_workbook(tickers, workbook):
    sheet = workbook['Stocks']
    for index, ticker in enumerate(tickers):
        sheet.cell(row = index + 2, column = 1).value = ticker
        
    workbook.save(save_name)

config_name = "config.yaml"
template_name = "stock_archive_template.xlsx"
save_name = "stock_archive_" + datetime.now().strftime("%Y%m%d-%H%M%S") + ".xlsx"
tmp_save_folder = "tmp/"
headers = ["timestamp", "open", "high", "low", "close", "adjusted_close", "volume", "dividend_amount", "split_coefficient"]

config = load_yaml(config_name)
workbook = load_workbook(template_name)
valid_tickers = []
for ticker in config['tickers']:
    data_save_path = tmp_save_folder + ticker + '.csv'
    try:
        json_data = gather_api_stock_data(ticker, config['api_key'])
        write_data(data_save_path, json_data, headers)   
        write_stock_data_to_workbook(data_save_path, workbook, ticker)           
        workbook.save(save_name)
        valid_tickers.append(ticker)
    except:
        print("No data found for " + ticker + ", skipping this one")

write_stock_names_to_workbook(valid_tickers, workbook)    
clear_folder(tmp_save_folder)

if config['email'] is not None:
    emailer(config['message'], config['subject'], [save_name], config['recipients'], config['email'], config['email_pass'])       